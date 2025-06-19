from openpyxl import load_workbook
import pyperclip
import tkinter as tk
import tkinter.font as font
from config import FILEPATH,TXT_PATH,WIDTH,HEIGHT,B_WIDTH,B_HEIGHT,X_COPY,Y_COPY,FONT_SIZE,X_QUIT,Y_QUIT

def road_txt():
    try:
        with open(TXT_PATH, 'r') as f:
            return int(f.read().strip())
    except(FileNotFoundError, ValueError):
        return 1
    
def save_txt(num):
    with open(TXT_PATH, 'w') as f:
        f.write(str(num))

#GUI生成クラス
class GUIApp():
    def __init__(self):
        self.num = road_txt()
        #print(f'{TXT_PATH}を読み込みました')
        #print(f'次の読み込みは{self.num}行目')
        self.wb = load_workbook(filename=FILEPATH)
        self.ws = self.wb.active

        self.root = tk.Tk()
        w=WIDTH
        h=HEIGHT
        self.root.geometry(f'{w}x{h}')
        self.root.update_idletasks()
        self.screen_width = self.root.winfo_width()#ウィンドウの幅を取得
        self.screen_height = self.root.winfo_height()#ウィンドウの高さを取得
        self.canvas = tk.Canvas(self.root, width=self.screen_width, height=self.screen_height)
        self.canvas.pack(fill="both", expand=True)

        self.text_1_min_xy = self.calc_area(4/5,3/20)
        self.log_text = f'読み込んだファイル : {FILEPATH}'
        self.text_1 = self.canvas.create_text(self.text_1_min_xy[0],self.text_1_min_xy[1],text=self.log_text)

        self.text_2_min_xy = self.calc_area(4/5,5/20)
        self.next_text = f'次の読み込みは{self.num}行目'
        self.text_2 = self.canvas.create_text(self.text_2_min_xy[0],self.text_2_min_xy[1],text=self.next_text)

        self.text_3_min_xy = self.calc_area(4/5,4/20)
        self.load_log_text = f'まだコピーしてません'
        self.text_3 = self.canvas.create_text(self.text_3_min_xy[0],self.text_3_min_xy[1],text=self.load_log_text)

        my_font = font.Font(family='Helvetica', size=FONT_SIZE)
        self.button = tk.Button(self.root, text = 'copy', command = self.click_copy, font=my_font)
        min_xy = self.calc_area(X_COPY, Y_COPY)
        self.button.place(x=min_xy[0],y=min_xy[1],width=B_WIDTH,height=B_HEIGHT)

        self.quit_button = tk.Button(self.root, text = "Quit", command=self.on_quit, font=my_font)
        min_xy = self.calc_area(X_QUIT, Y_QUIT)
        self.quit_button.place(x=min_xy[0],y=min_xy[1],width=B_WIDTH,height=B_HEIGHT)

    def on_quit(self):
        print('終了処理中...')
        save_txt(self.num)
        self.root.destroy()

    def calc_area(self, center_x, center_y):
        half_x = B_WIDTH // 2
        half_y = B_HEIGHT //2
        min_x = center_x * WIDTH - half_x
        min_y = center_y * HEIGHT - half_y
        return (min_x, min_y)

    def click_copy(self):
        self.val = self.ws[f'A{self.num}'].value
        pyperclip.copy(self.val)
        #print(f'A{self.num}:{self.val}をコピー')
        self.canvas.itemconfig(self.text_3, text=f'A{self.num}:{self.val}をコピー')
        self.canvas.itemconfig(self.text_2, text=f'次の読み込みは{self.num+1}行目')
        self.num = self.num + 1

    def run(self):
        self.root.mainloop()

def main():
    app = GUIApp()
    app.run()
if __name__ == '__main__':
    main()
